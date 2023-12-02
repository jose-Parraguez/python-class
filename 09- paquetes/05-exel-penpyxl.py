import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\Parraguez\\Desktop\\fruta.xlsx")

print(wb["Hoja1"])

hoja = wb.active

# print(
#     hoja.max_row,
#     hoja.max_column,
# )

nombres_columnas = [celda.value for celda in hoja[3]]

# Imprimir los nombres de las columnas
print("Nombres de las columnas:", nombres_columnas)

for fila in range(1,hoja.max_row + 1):
    for columna in range(1,hoja.max_column + 1):
        celda = hoja.cell(row= fila, column= columna)
        print(celda.value)